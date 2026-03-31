import os
import requests
import datetime
import json
import re
import pandas as pd
import subprocess
import markdown
from google import genai
from dotenv import load_dotenv
from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# === 基礎配置 ===
load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
SERPER_API_KEY = os.getenv("SERPER_API_KEY")
LINE_CHANNEL_ACCESS_TOKEN = os.getenv("LINE_CHANNEL_ACCESS_TOKEN")
LINE_USER_ID = os.getenv("LINE_USER_ID")
REPORT_DIR = "docs/reports"
GITHUB_IO_URL = "https://aphsu124.github.io/myAgent"
ICLOUD_BASE = "/Users/bucksteam/Library/Mobile Documents/com~apple~CloudDocs/泰國/工作/甲米油廠/簡報"
ICLOUD_EXCEL = os.path.join(ICLOUD_BASE, "palm_oil_history.xlsx")

# 註冊中文字體 (PDF 專用)
pdfmetrics.registerFont(TTFont('Chinese', '/System/Library/Fonts/STHeiti Light.ttc'))
client = genai.Client(api_key=GEMINI_API_KEY, http_options={'api_version': 'v1'})

# --- 1. 核心修復：Excel 專業美化 (完全還原老闆最滿意版本) ---
def style_excel(file_path):
    try:
        wb = load_workbook(file_path); ws = wb.active
        # 1.1 解除所有合併 (防止報錯)
        for m_range in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m_range))
        
        # 1.2 定義樣式 (14號字，微軟正黑體)
        f_name = 'Microsoft JhengHei'
        header_f = Font(name=f_name, size=14, bold=True, color="FFFFFF")
        data_f = Font(name=f_name, size=14)
        guide_title_f = Font(name=f_name, size=14, bold=True, color="FFFFFF")
        guide_text_f = Font(name=f_name, size=13, color="333333")
        
        center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
        left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        # 1.3 美化數據區 (A-G)
        for c in range(1, 8):
            cell = ws.cell(row=1, column=c)
            cell.font = header_f; cell.fill = PatternFill(start_color="2E7D32", end_color="2E7D32", fill_type="solid")
            cell.alignment = center_align; cell.border = border
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=c)
                cell.font = data_f; cell.alignment = center_align; cell.border = border
                if isinstance(cell.value, (int, float)): cell.number_format = '0.00'
        
        # 1.4 建立指南區 (I-J)
        guide_col = 9
        ws.cell(row=1, column=guide_col).value = "📚 產業術語與計算指南 (Guide)"
        ws.cell(row=1, column=guide_col).font = guide_title_f
        ws.cell(row=1, column=guide_col).fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
        ws.cell(row=1, column=guide_col).alignment = center_align
        ws.merge_cells(start_row=1, start_column=guide_col, end_row=1, end_column=guide_col+1)
        
        items = [
            ["FFB", "Fresh Fruit Bunch : 油棕鮮果串。工廠收購的原始原料。"],
            ["CPO", "Crude Palm Oil : 毛棕櫚油。工廠壓榨產出的核心產品。"],
            ["BMD", "Bursa Malaysia Derivatives : 馬來西亞期貨。全球報價基準。"],
            ["Basis", "基差 = 泰國CPO - (BMD/1000 * 匯率)。反映國內外溢價。"],
            ["EX Rate", "Exchange Rate : 馬幣(MYR) 兌 泰銖(THB) 匯率。"]
        ]
        for i, item in enumerate(items):
            ws.cell(row=i+2, column=guide_col).value = item[0]; ws.cell(row=i+2, column=guide_col).font = Font(name=f_name, bold=True, size=14)
            ws.cell(row=i+2, column=guide_col+1).value = item[1]; ws.cell(row=i+2, column=guide_col+1).font = guide_text_f; ws.cell(row=i+2, column=guide_col+1).alignment = left_align
        
        # 1.5 精確設定欄寬
        widths = {'A': 22, 'B': 16, 'C': 16, 'D': 16, 'E': 16, 'F': 22, 'G': 16, 'H': 5, 'I': 20, 'J': 85}
        for l, w in widths.items(): ws.column_dimensions[l].width = w
        
        wb.save(file_path)
        print("✅ Excel 專業樣式還原成功！")
    except Exception as e: print(f"Excel 失敗: {e}")

# --- 2. PDF 專業排版 (ReportLab) ---
def create_pdf_report(filename, title, date, ffb, cpo, content):
    try:
        doc = SimpleDocTemplate(filename, pagesize=A4)
        p_style = ParagraphStyle('CN', fontName='Chinese', fontSize=12, leading=18, spaceAfter=10)
        t_style = ParagraphStyle('Title', fontName='Chinese', fontSize=18, leading=24, alignment=1, spaceAfter=20)
        story = [Paragraph(title, t_style), Paragraph(f"日期: {date} | FFB: {ffb} | CPO: {cpo}", p_style), Spacer(1, 12)]
        clean = content.replace('#', '').replace('*', '')
        for line in clean.split('\n'):
            if line.strip(): story.append(Paragraph(line, p_style))
        doc.build(story)
    except: pass

# --- 3. 數據更新邏輯 ---
def update_icloud_excel(date_str, ffb, cpo, bmd_myr, ex_rate):
    try:
        def to_f(v):
            try: return float(v)
            except: return 0.0
        f_ffb, f_cpo, f_bmd, f_ex = to_f(ffb), to_f(cpo), to_f(bmd_myr), to_f(ex_rate)
        bmd_thb = round((f_bmd / 1000) * f_ex, 2); basis = round(f_cpo - bmd_thb, 2)
        new_r = pd.DataFrame({"Date": [date_str], "FFB": [f_ffb], "CPO": [f_cpo], "BMD_MYR": [f_bmd], "Exchange_Rate": [f_ex], "BMD_THB_kg": [bmd_thb], "Basis": [basis]})
        if os.path.exists(ICLOUD_EXCEL):
            df_old = pd.read_excel(ICLOUD_EXCEL)
            if date_str in df_old['Date'].values:
                for col in new_r.columns: df_old.loc[df_old['Date'] == date_str, col] = new_r[col].values[0]
                df_old.to_excel(ICLOUD_EXCEL, index=False)
            else: pd.concat([df_old, new_r], ignore_index=True).to_excel(ICLOUD_EXCEL, index=False)
        else: new_r.to_excel(ICLOUD_EXCEL, index=False)
        style_excel(ICLOUD_EXCEL)
        return bmd_thb, basis
    except: return 0, 0

def main():
    ict_now = datetime.datetime.now(datetime.timezone(datetime.timedelta(hours=7)))
    date_fn, date_ds = ict_now.strftime("%Y%m%d"), ict_now.strftime("%Y-%m-%d")
    curr_hm = ict_now.strftime("%H:%M")
    
    if "07:00" <= curr_hm < "13:30":
        mode, title, suffix = "news_only", "📰 泰國棕櫚油晨間新聞", "M_report"
    elif curr_hm >= "13:30":
        mode, title, suffix = "full", "📊 泰國棕櫚油完整報告", "D_report"
    else: return

    print(f"🚀 [全能完美版] 啟動 {mode} 任務...")
    
    # 抓取、分析、同步 (簡化省略，同之前邏輯)
    url = "https://google.serper.dev/search"; res = ""
    for q in [f"Thailand palm oil price {date_ds} -travel", f"Bursa Malaysia CPO futures {date_ds}"]:
        try:
            r = requests.post(url, headers={'X-API-KEY': SERPER_API_KEY}, json={"q": q, "gl": "th", "hl": "en", "tbs": "qdr:m"})
            if r.status_code == 200:
                for o in r.json().get('organic', []): res += f"\n{o.get('snippet')}\n"
        except: pass

    content, data = "無法生成", {"ffb": "N/A", "cpo": "N/A", "bmd_myr": "N/A", "ex_rate": "N/A"}
    try:
        resp = client.models.generate_content(model="gemini-2.5-flash", contents=f"撰寫繁體中文報告並輸出 JSON: DATA_JSON: {{...}} \n 數據：{res}")
        if resp.text:
            content = resp.text
            m = re.search(r'DATA_JSON: ({.*?})', content)
            if m: data = json.loads(m.group(1))
    except: pass

    bmd_thb, basis = update_icloud_excel(date_ds, data.get('ffb'), data.get('cpo'), data.get('bmd_myr'), data.get('ex_rate'))
    create_pdf_report(os.path.join(ICLOUD_BASE, f"{date_fn}_{suffix}.pdf"), title, date_ds, data.get('ffb'), data.get('cpo'), content)

    # 網頁
    html_body = markdown.markdown(content)
    web_content = f"<html><body style='background:#121212;color:#e0e0e0;padding:40px;font-family:sans-serif;'><h1>{title}</h1>{html_body}</body></html>"
    with open(os.path.join(REPORT_DIR, f"{date_fn}_{suffix}.html"), "w", encoding="utf-8") as f: f.write(web_content)
    with open("docs/index.html", "w", encoding="utf-8") as f: f.write(web_content)
    
    try:
        subprocess.run(["git", "add", "."], check=True)
        subprocess.run(["git", "commit", "-m", f"📊 Stable Update {date_fn}"], check=True)
        subprocess.run(["git", "push", "origin", "main"], check=True)
    except: pass

    msg = f"{title}\n🔸 FFB: {data.get('ffb')}\n🔸 CPO: {data.get('cpo')}\n🔸 基差: {basis}\n👉 查看網頁：{GITHUB_IO_URL}/index.html"
    requests.post("https://api.line.me/v2/bot/message/push", headers={"Content-Type": "application/json", "Authorization": f"Bearer {LINE_CHANNEL_ACCESS_TOKEN}"}, json={"to": LINE_USER_ID, "messages": [{"type": "text", "text": msg}]})

if __name__ == "__main__":
    main()
