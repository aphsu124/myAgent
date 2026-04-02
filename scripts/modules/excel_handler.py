import os
import io
import pandas as pd
import matplotlib.pyplot as plt
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from .config import (
    ICLOUD_EXCEL, MS_JH_FONT, ICLOUD_CHART,
    STORAGE_BACKEND, GDRIVE_FOLDER_BRIEFING, GDRIVE_EXCEL_FILE_ID
)

# 設置中文字體 (針對 Mac)
plt.rcParams['font.sans-serif'] = ['Arial Unicode MS']
plt.rcParams['axes.unicode_minus'] = False

def generate_trend_chart():
    """繪製最新棕櫚油趨勢圖"""
    try:
        # 讀取數據
        if STORAGE_BACKEND == 'gdrive' and GDRIVE_EXCEL_FILE_ID:
            from .gdrive_utils import read_excel_to_dataframe, upload_bytes
            df = read_excel_to_dataframe(GDRIVE_EXCEL_FILE_ID)
        else:
            if not os.path.exists(ICLOUD_EXCEL): return
            df = pd.read_excel(ICLOUD_EXCEL)

        if df is None or len(df) < 2: return

        plt.figure(figsize=(10, 6))
        plt.plot(df['Date'], df['FFB'], label='FFB 現貨', marker='o', color='green')
        plt.plot(df['Date'], df['CPO'], label='CPO 現貨', marker='s', color='orange')
        plt.plot(df['Date'], df['BMD_THB_kg'], label='BMD 期貨 (折算泰銖)', linestyle='--', color='blue')

        plt.title('2026 泰國棕櫚油價格趨勢圖', fontsize=16)
        plt.xlabel('日期', fontsize=12); plt.ylabel('泰銖 (THB/kg)', fontsize=12)
        plt.legend(); plt.grid(True, linestyle=':', alpha=0.6)
        plt.xticks(rotation=45)
        plt.tight_layout()

        if STORAGE_BACKEND == 'gdrive' and GDRIVE_FOLDER_BRIEFING:
            buf = io.BytesIO()
            plt.savefig(buf, format='png')
            plt.close()
            buf.seek(0)
            upload_bytes(buf.read(), "2026_Palm_Oil_Trend_Master.png", GDRIVE_FOLDER_BRIEFING, "image/png")
            print("📈 趨勢圖已上傳至 Google Drive")
        else:
            plt.savefig(ICLOUD_CHART)
            plt.close()
            print(f"📈 趨勢圖已更新: {ICLOUD_CHART}")
    except Exception as e:
        print(f"❌ 繪圖失敗: {e}")

def apply_excel_style_to_wb(wb):
    """套用 Excel 樣式至 openpyxl Workbook 物件（不含磁碟 I/O）"""
    try:
        ws = wb.active
        for m_range in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m_range))

        h_f = Font(name=MS_JH_FONT, size=14, bold=True, color="FFFFFF")
        d_f = Font(name=MS_JH_FONT, size=14)
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))

        for c in range(1, 8):
            cell = ws.cell(row=1, column=c)
            cell.font = h_f; cell.fill = PatternFill('solid', fgColor='2E7D32'); cell.alignment = Alignment('center', 'center')
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(row=r, column=c); cell.font = d_f; cell.border = border; cell.alignment = Alignment('center', 'center')
                if isinstance(cell.value, (int, float)): cell.number_format = '0.00'

        g_c = 9; ws.cell(1, g_c).value = "📚 產業術語與計算指南"
        ws.cell(1, g_c).font = h_f; ws.cell(1, g_c).fill = PatternFill('solid', fgColor='1565C0')
        ws.merge_cells(start_row=1, start_column=g_c, end_row=1, end_column=g_c+1)
        ws.cell(1, g_c).alignment = Alignment('center', 'center')

        items = [
            ["FFB", "Fresh Fruit Bunch : 油棕鮮果串。工廠收購的原始原料。"],
            ["CPO", "Crude Palm Oil : 毛棕櫚油。工廠壓榨產出的核心產品。"],
            ["BMD_MYR", "馬來西亞期貨收盤價 (馬幣/公噸)。全球定價基準。"],
            ["BMD_THB_kg", "國際期貨折算價 (泰銖/公斤)。算式：(BMD_MYR/1000) * 匯率。"],
            ["Basis", "基差 = 泰國CPO - 國際期貨折算價。反映國內外價差。"],
            ["EX Rate", "Exchange Rate : 馬幣兌泰銖當日匯率。"]
        ]
        for i, item in enumerate(items):
            ws.cell(i+2, g_c).value = item[0]; ws.cell(i+2, g_c).font = Font(name=MS_JH_FONT, bold=True, size=14)
            ws.cell(i+2, g_c+1).value = item[1]; ws.cell(i+2, g_c+1).font = Font(name=MS_JH_FONT, size=14); ws.cell(i+2, g_c+1).alignment = Alignment(horizontal='left', wrap_text=True)

        ws.column_dimensions['A'].width = 25; ws.column_dimensions['D'].width = 22; ws.column_dimensions['E'].width = 22; ws.column_dimensions['J'].width = 110
    except Exception as e:
        print(f"Excel 樣式應用失敗: {e}")

def apply_excel_style():
    """舊版磁碟模式樣式套用（iCloud fallback 用）"""
    if not os.path.exists(ICLOUD_EXCEL): return
    try:
        wb = load_workbook(ICLOUD_EXCEL)
        apply_excel_style_to_wb(wb)
        wb.save(ICLOUD_EXCEL)
    except Exception as e:
        print(f"Excel 樣式應用失敗: {e}")

def update_data(date_str, ffb, cpo, bmd_myr, ex_rate):
    """更新數據並回傳計算結果"""
    try:
        f_ffb, f_cpo, f_bmd, f_ex = map(lambda x: float(x) if str(x).replace('.','').replace('-','').isdigit() else 0.0, [ffb, cpo, bmd_myr, ex_rate])
        bmd_thb = round((f_bmd / 1000) * f_ex, 2)
        basis = round(f_cpo - bmd_thb, 2) if f_cpo > 0 else 0.0
    except Exception as e:
        print(f"數據計算失敗: {e}")
        return 0, 0

    try:
        new_row = pd.DataFrame({"Date": [date_str], "FFB": [f_ffb], "CPO": [f_cpo], "BMD_MYR": [f_bmd], "Exchange_Rate": [f_ex], "BMD_THB_kg": [bmd_thb], "Basis": [basis]})

        if STORAGE_BACKEND == 'gdrive' and GDRIVE_EXCEL_FILE_ID:
            from .gdrive_utils import read_excel_to_dataframe, write_dataframe_to_excel
            df = read_excel_to_dataframe(GDRIVE_EXCEL_FILE_ID)
            if df is not None:
                if date_str in df['Date'].values:
                    for col in new_row.columns: df.loc[df['Date'] == date_str, col] = new_row[col].values[0]
                else:
                    df = pd.concat([df, new_row], ignore_index=True)
            else:
                df = new_row
            write_dataframe_to_excel(df, GDRIVE_EXCEL_FILE_ID, apply_style_fn=apply_excel_style_to_wb)
        else:
            if os.path.exists(ICLOUD_EXCEL):
                df = pd.read_excel(ICLOUD_EXCEL)
                if date_str in df['Date'].values:
                    for col in new_row.columns: df.loc[df['Date'] == date_str, col] = new_row[col].values[0]
                else:
                    df = pd.concat([df, new_row], ignore_index=True)
                df.to_excel(ICLOUD_EXCEL, index=False)
            else:
                new_row.to_excel(ICLOUD_EXCEL, index=False)
            apply_excel_style()

    except Exception as e:
        print(f"Excel 檔案更新失敗 (可能是權限問題): {e}")

    return bmd_thb, basis
