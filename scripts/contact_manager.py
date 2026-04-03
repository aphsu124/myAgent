import os
import json
import re
import io
import pandas as pd
import fitz
from PIL import Image
from google import genai
from dotenv import load_dotenv
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

load_dotenv()
GEMINI_API_KEY = os.getenv("GEMINI_API_KEY")
from modules.config import (
    STORAGE_BACKEND, GDRIVE_FOLDER_CONTACT,
    GDRIVE_CONTACT_FILE_ID, GDRIVE_CONTACT_SHEET_GID, MS_JH_FONT
)

client = genai.Client(api_key=GEMINI_API_KEY, http_options={'api_version': 'v1'})

def clean_str(s):
    return str(s).strip().lower() if s else ""

def _open_as_image(file_path):
    """將檔案（圖片或 PDF）統一轉為 PIL Image（取第一頁）"""
    if file_path.lower().endswith('.pdf'):
        doc = fitz.open(file_path)
        pix = doc[0].get_pixmap(matrix=fitz.Matrix(2, 2))
        img = Image.open(io.BytesIO(pix.tobytes("png")))
        doc.close()
        return img
    else:
        return Image.open(file_path)

def style_excel_wb(wb):
    """套用樣式至 openpyxl Workbook 物件（iCloud fallback 用）"""
    try:
        ws = wb.active
        for m_range in list(ws.merged_cells.ranges): ws.unmerge_cells(str(m_range))
        h_f = Font(name=MS_JH_FONT, size=14, bold=True, color="FFFFFF")
        d_f = Font(name=MS_JH_FONT, size=14)
        h_fill = PatternFill('solid', fgColor='1565C0')
        border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(1, c)
            cell.font = h_f; cell.fill = h_fill
            cell.alignment = Alignment('center', 'center', wrap_text=True); cell.border = border
            ws.column_dimensions[cell.column_letter].width = 45 if cell.value in ['備註', '地址'] else 30
            for r in range(2, ws.max_row + 1):
                cell = ws.cell(r, c); cell.font = d_f; cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    except Exception as e:
        print(f"⚠️ 聯絡人 Excel 樣式失敗: {e}")

def save_to_excel_smart(data):
    """極致去重且支援新欄位：Drive 模式寫 Google Sheets，否則寫本機 xlsx"""
    c_name_cn = clean_str(data.get('姓名(中)'))
    c_name_en = clean_str(data.get('姓名(英)'))
    c_comp = clean_str(data.get('公司'))

    if STORAGE_BACKEND == 'gdrive' and GDRIVE_CONTACT_FILE_ID:
        from modules.gdrive_utils import (
            read_sheet_to_dataframe, append_row_to_sheet, update_row_in_sheet
        )
        df = read_sheet_to_dataframe(GDRIVE_CONTACT_FILE_ID, GDRIVE_CONTACT_SHEET_GID)
        if df is not None and not df.empty:
            df = df.fillna("")
            for col in ['姓名(中)', '姓名(英)', '公司']:
                if col not in df.columns: df[col] = ""
            mask = (
                ((df['姓名(中)'].apply(clean_str) == c_name_cn) & (c_name_cn != "")) |
                ((df['姓名(英)'].apply(clean_str) == c_name_en) & (c_name_en != ""))
            ) & (df['公司'].apply(clean_str) == c_comp)
            match = df[mask]
            if not match.empty:
                # 只更新那一行，其他資料完全不動
                print(f"🔄 更新既有資料: {data.get('姓名(中)') or data.get('姓名(英)')}")
                idx = match.index[0]
                for key, val in data.items():
                    if val: df.at[idx, key] = val
                row_values = [str(df.at[idx, col]) for col in df.columns]
                update_row_in_sheet(idx, row_values, GDRIVE_CONTACT_FILE_ID, GDRIVE_CONTACT_SHEET_GID)
            else:
                # 只在底部附加一行，不動現有資料
                print(f"➕ 新增聯絡人: {data.get('姓名(中)') or data.get('姓名(英)')}")
                row_values = [str(data.get(col, '')) for col in df.columns]
                append_row_to_sheet(row_values, GDRIVE_CONTACT_FILE_ID, GDRIVE_CONTACT_SHEET_GID)
        else:
            # 表格為空：直接附加（標題列已由使用者建立）
            print(f"➕ 新增第一筆聯絡人: {data.get('姓名(中)') or data.get('姓名(英)')}")
            row_values = [str(data.get(col, '')) for col in ['姓名(中)', '姓名(英)', '公司', '職稱', '稅號', '網址', '電話', 'Email', '地址', '備註']]
            append_row_to_sheet(row_values, GDRIVE_CONTACT_FILE_ID, GDRIVE_CONTACT_SHEET_GID)
    else:
        excel_path = "/Users/bucksteam/Library/Mobile Documents/com~apple~CloudDocs/泰國/工作/聯絡人/人脈總表.xlsx"
        if os.path.exists(excel_path):
            df = pd.read_excel(excel_path).fillna("")
            mask = (
                ((df['姓名(中)'].apply(clean_str) == c_name_cn) & (c_name_cn != "")) |
                ((df['姓名(英)'].apply(clean_str) == c_name_en) & (c_name_en != ""))
            ) & (df['公司'].apply(clean_str) == c_comp)
            match = df[mask]
            if not match.empty:
                idx = match.index[0]
                for key, val in data.items():
                    if val: df.at[idx, key] = val
                df.to_excel(excel_path, index=False)
            else:
                pd.concat([df, pd.DataFrame([data])], ignore_index=True).to_excel(excel_path, index=False)
        else:
            pd.DataFrame([data]).to_excel(excel_path, index=False)
        try:
            from openpyxl import load_workbook
            wb = load_workbook(excel_path)
            style_excel_wb(wb)
            wb.save(excel_path)
        except Exception as e:
            print(f"⚠️ 樣式套用失敗: {e}")

def process_contact_file(file_path, file_id=None):
    fn = os.path.basename(file_path)
    ext = os.path.splitext(fn)[1].lower()
    print(f"📇 執行精確分類辨識: {fn}")

    prompt = """
    任務：提取聯絡資訊。
    【欄位規範】：
    1. 姓名：拆分 [姓名(中)] 與 [姓名(英)]。
    2. 稅號：提取統一編號或 Tax ID。
    3. 網址：提取公司官網。
    4. 電話：格式 [手機: xxx\\n公司: xxx]。
    5. 備註：類別化，且每一類之間必須加入「兩個換行符號」以利閱讀。範例:「類別: 內容\\n\\n類別: 內容」。

    JSON: {"姓名(中)": "", "姓名(英)": "", "公司": "", "職稱": "", "稅號": "", "網址": "", "電話": "", "Email": "", "地址": "", "備註": ""}
    """

    try:
        img = _open_as_image(file_path)
        resp = client.models.generate_content(model="gemini-2.5-flash", contents=[prompt, img])
        try:
            from modules.token_tracker import record as _tt
            _tt('google', 'gemini-2.5-flash', resp.usage_metadata.prompt_token_count or 0, resp.usage_metadata.candidates_token_count or 0)
        except Exception: pass
        m = re.search(r'({.*?})', resp.text, re.DOTALL)
        if m:
            data = json.loads(m.group(1))
            save_to_excel_smart(data)
            if STORAGE_BACKEND == 'gdrive' and file_id:
                from modules.gdrive_utils import delete_file
                delete_file(file_id)
                print(f"✅ 完成，來源檔已從 Drive 刪除：{fn}")
            else:
                base_dir = "/Users/bucksteam/Library/Mobile Documents/com~apple~CloudDocs/泰國/工作/聯絡人"
                proc_dir = os.path.join(base_dir, "processed")
                if not os.path.exists(proc_dir): os.makedirs(proc_dir)
                import shutil
                shutil.move(file_path, os.path.join(proc_dir, fn))
                print(f"✅ 完成，來源檔已歸檔")
        else:
            print(f"⚠️ AI 未回傳有效 JSON，跳過：{fn}")
    except Exception as e:
        print(f"❌ 錯誤: {e}")

def main():
    if STORAGE_BACKEND == 'gdrive' and GDRIVE_FOLDER_CONTACT:
        from modules.gdrive_utils import list_files_in_folder, download_file
        supported_exts = ('.jpg', '.jpeg', '.png', '.bmp', '.webp', '.pdf')
        files = list_files_in_folder(GDRIVE_FOLDER_CONTACT)
        files = [f for f in files if any(f['name'].lower().endswith(ext) for ext in supported_exts)]
        if not files:
            print("📂 Drive 聯絡人資料夾中無待處理名片（支援 jpg/png/pdf）。")
            return
        for f in files:
            tmp_path = f"/tmp/{f['name']}"
            if download_file(f['id'], tmp_path):
                process_contact_file(tmp_path, file_id=f['id'])
                try: os.remove(tmp_path)
                except: pass
    else:
        base_dir = "/Users/bucksteam/Library/Mobile Documents/com~apple~CloudDocs/泰國/工作/聯絡人"
        supported_exts = ('.jpg', '.jpeg', '.png', '.bmp', '.webp', '.pdf')
        files = [
            os.path.join(base_dir, f) for f in os.listdir(base_dir)
            if os.path.isfile(os.path.join(base_dir, f))
            and not f.startswith(".")
            and os.path.splitext(f)[1].lower() in supported_exts
        ]
        for f in files:
            process_contact_file(f)

if __name__ == "__main__":
    main()
